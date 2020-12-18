from openpyxl import load_workbook
from openpyxl.comments import Comment

# open excel file
wb = load_workbook(filename='ExcelData01_Test.xlsx')

# get a worksheet
ws = wb['Sheet1']

# process 4 rows for the example
# this is one more than there are rows with comments
# we start the processing on row 3, because that's where the data starts
start_row = 3
num_rows = 4
column_letter_to_read = 'A'  # column A contains the comment to be copied to another cell value
column_letter_to_write = 'B'  # column B will get the comment in column A as a value and also a comment
comment_author = 'Brian'
comment_count = 0
for i in range(start_row, start_row + num_rows + 1):  # the '+ 1' is needed because Python range() "stop" parameter is exclusive
    read_cell_id = column_letter_to_read + str(i)
    # print(ws[ref].value)
    comment = ws[read_cell_id].comment
    if comment:
        comment_count += 1
        # print(f"{read_cell_id}: {comment.text}")
        write_cell_id = column_letter_to_write + str(i)
        ws[write_cell_id].value = comment.text
        ws[write_cell_id].comment = Comment(comment.text, comment_author)

wb.save(filename='ExcelData01_Output.xlsx')
wb.close()
print(f"Successfully processed {comment_count} comments.")
