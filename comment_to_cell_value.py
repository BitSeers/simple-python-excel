# consult https://openpyxl.readthedocs.io/en/stable/ for more info
from openpyxl import load_workbook

# input and output files
OUTPUT_FILENAME = 'ExcelData01_Output.xlsx'
INPUT_FILENAME = 'ExcelData01_Test.xlsx'

# open excel file
wb = load_workbook(filename=INPUT_FILENAME)

# list of worksheets to process
ws_list = [f'{i}.0' for i in range(1, 14)]

# get a worksheet
ws = wb[ws_list[0]]

comment_column_letter = 'C'
desc_column_letter = 'D'
comment_counter = 0

# loop through all the worksheets that need comment processing
for ws_name in ws_list:
    ws = wb[ws_name]
    for t in range(1, 1000):
        read_cell_name = comment_column_letter + str(t)
        # print(ws[ref].value)  # DEBUG HELPER
        comment = ws[read_cell_name].comment
        # if a comment exists then write it to the corresponding cell as a value
        if comment:
            comment_counter += 1
            # print(f"{read_cell_name}: {comment.text}")  # DEBUG HELPER
            write_cell_name = desc_column_letter + str(t)
            ws[write_cell_name].value = comment.text

wb.save(filename=OUTPUT_FILENAME)
wb.close()
print(f"Successfully processed {comment_counter} comments.")
